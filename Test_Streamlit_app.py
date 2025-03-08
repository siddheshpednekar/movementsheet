import streamlit as st
import pandas as pd
from io import BytesIO
import re
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook, load_workbook


# Streamlit App
st.title("Movement Sheet Web App")

# File uploader for Excel files
package_booking = st.file_uploader("Upload Package Booking file", type=["xlsx", "xls"])
standalone = st.file_uploader("Upload Standalone file", type=["xlsx", "xls"])
package_file_path = r"Package File Movement sheet.xlsx"

check_df = None
check_df2 = None
if st.button('Start Processing'):
    # Read the Excel file into a DataFrame
        
    if package_booking is None:
        st.write("Package booking file not uploaded")

    elif package_booking is not None:
        df = pd.read_excel(package_booking)
        required_columns = {"BookingStatus", "ServiceType", "ServiceTime", "ETA", "DepartureFlightNumber",
                                "ArrivalFlightNumber", "TransitFlightNumber", "ETD", "Origin", "Destination",
                                "PackageName", "Nationality", "TravelClass", "Remarks"}
        # Check for missing columns
        missing_columns = required_columns - set(df.columns)
        if missing_columns:
            st.error(f"Missing columns in Package Booking file: {', '.join(missing_columns)}")
        else:
            df = pd.read_excel(package_booking)
            check_df = df
            package_df = pd.read_excel(package_file_path)
            file1 = "read"
    
            
    
            
        
            st.write("### Uploaded Package booking file:")
            st.dataframe(df)
        
            # Package Mapping
            package_mapping = dict(zip(package_df["Package_Name"], package_df["Revised_Package_Name"]))
        
        
            # Step 2: Filter BookingStatus
            df = df[df["BookingStatus"].isin(["Completed", "PaymentCompleted"])]
        
            # Step 3: Adjust "ServiceType" for RoundTrip
            df.loc[(df["ServiceType"] == "RoundTrip") & (df["ServiceTime"] == df["ETA"]), "ServiceType"] = "Arrival"
            df.loc[(df["ServiceType"] == "RoundTrip") & (df["ServiceTime"] != df["ETA"]), "ServiceType"] = "Departure"
        
            # Step 4: Create "Flight No." column
            def get_flight_no(row):
                if row["ServiceType"] == "Departure" and row["DepartureFlightNumber"] not in ["NA", "", None]:
                    return row["DepartureFlightNumber"]
                elif row["ServiceType"] == "Arrival" and row["ArrivalFlightNumber"] not in ["NA", "", None]:
                    return row["ArrivalFlightNumber"]
                elif row["ServiceType"] == "Transit":
                    return f"{row['ArrivalFlightNumber']} / {row['TransitFlightNumber']}"
                return ""
        
            df["Flight No."] = df.apply(get_flight_no, axis=1)
        
            # Step 5: Create "ETA/ETD" column
            df["ETA/ETD"] = df["ETD"].fillna(df["ETA"])
        
            # Step 6: Clean "Origin" and "Destination" columns
            def clean_location(value):
                return "" if pd.isna(value) else re.split(r"[,/]", value)[0].strip()
        
            df["Origin"] = df["Origin"].apply(clean_location)
            df["Destination"] = df["Destination"].apply(clean_location)
        
            # Step 7: Create "Orig/Dest" column
            def get_orig_dest(row):
                if row["ServiceType"] == "Arrival":
                    return row["Origin"]
                elif row["ServiceType"] in ["Departure", "RoundTrip"]:
                    return row["Destination"]
                elif row["ServiceType"] == "Transit":
                    return f"{row['Origin']} / {row['Destination']}"
                return ""
        
            df["Orig/Dest"] = df.apply(get_orig_dest, axis=1)
        
            # Step 8: Create "Terminal" column
            def get_terminal(flight_no):
                match = re.search(r"-(\d{4})", str(flight_no))
                return "T1" if match and match.group(1).startswith("5") else "T2"
        
            df["Terminal"] = df["Flight No."].apply(get_terminal)
        
            # Step 9: Map "Package" column
            df["Package"] = df["PackageName"].apply(lambda x: package_mapping.get(x, x))
        
            # Step 10: Add empty "Profile" and "GSO" columns
            df["Profile"] = ""
            df["GSO"] = ""
        
            # Step 11: Create "Serial Number" column
            df.insert(0, "Sr No.", range(1, len(df) + 1))
        
            # Step 12: Remove duplicate values in "Nationality" and "Class of Travel"
            def remove_duplicates(text):
                return "" if pd.isna(text) else " : ".join(dict.fromkeys(text.split(" : ")))
        
            df["Nationality"] = df["Nationality"].apply(remove_duplicates)
            df["TravelClass"] = df["TravelClass"].apply(remove_duplicates)
        
            # Step 13: Rename columns
            rename_columns = {
           "Serial Number": "Sr No.", "ServiceTime": "Service Time", "ItenaryNumber": "Itinerary No.",
            "ServiceType": "Service", "Package": "Package", "Terminal": "Terminal",
            "GuestName": "GUEST NAME", "TotalGuest": "Total Guest", "Flight No.": "Flight No.",
            "Orig/Dest": "Orig/Dest", "ETA/ETD": "ETA/ETD", "TravelClass": "Class of Travel",
            "PlacardCountryCode": "Country Code", "PlacardContactNo": "Placard Guest Contact No.",
            "PlacardName": "Placard Guest Name", "Nationality": "Nationality", "Age": "Age",
            "BillingContactNo": "Booker Contact No.", "BillingEmail": "Email Id", "Remarks": "Remarks"
            }
            df = df.rename(columns=rename_columns)
        
            # Step 14: Remove rows where "Remarks" is "Cancelled"
            df = df[df["Remarks"] != "Cancelled"]
        
            # Step 15: Reorder columns
            column_order = ["Sr No.", "Service Time", "Itinerary No.", "Service", "Package", "Terminal", "Profile",
                            "GUEST NAME", "Total Guest", "Flight No.", "Orig/Dest", "ETA/ETD", "Class of Travel",
                            "Country Code", "Placard Guest Contact No.", "Placard Guest Name", "Nationality", "Age",
                            "Booker Contact No.", "Email Id", "Remarks", "GSO"]
            #df = df[column_order]
            
            # Check which columns are missing
            missing_cols = [col for col in column_order if col not in df.columns]
            if missing_cols:
                st.write("⚠️ Missing columns:", missing_cols)
        
            # Reorder only existing columns
            df = df[[col for col in column_order if col in df.columns]]
        
            # Step 16: Save DataFrame to an Excel file in memory
            output = BytesIO()
        
            
            
            # Step 16: Save to Excel with Bold Formatting
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Sheet1")
                workbook = writer.book
                sheet = writer.sheets["Sheet1"]
                bold_font = Font(bold=True)
            
                # Apply Bold only on Package Column
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
                    for cell in row:
                        if str(cell.value).strip() in map(str.strip, package_mapping.values()):
                            cell.font = bold_font
            
            excel_df = output
            #st.write("✅ Data processing completed successfully!")
            #st.write("### Processed Data:")
            #st.write(df)
            
            # Step 17: Download Button
            #st.download_button(
                #label="Download Package Booking File",
                #data=output,
                #file_name="modified_data.xlsx",
                #mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            #)
    

    if standalone is None:
        st.write("Standalone file not uploaded")
    elif standalone is not None:
        df2 = pd.read_excel(standalone)
        # Required Columns
        required_columns = {
            "BookingStatus", "ServiceType", "ServiceTime", "ItenaryNumber", "Airline",
            "Origin", "Destination", "ServiceDetail", "TotalPorters", "BookingPersonName",
            "BookingPersonContactNo", "BookingPersonEmail", "Remarks", "GSO"
        }
    
        # Check Missing Columns
        missing_columns2 = required_columns - set(df2.columns)
        if missing_columns2:
            st.error(f"Missing columns in Standalone file: {', '.join(missing_columns2)}")
        
        else:
            df2 = pd.read_excel(standalone)
            check_df2 = df2
            file2 = "read"
            st.write("### Uploaded Standalone file:")
            st.dataframe(df2)
    
    
            # Step 1: Create "Sr No" column
            df2.insert(0, "Sr No", range(1, len(df2) + 1))
            
            # Step 2: Create "Terminal" column
            def get_terminal(airline):
                match = re.search(r"-(\d{4})", str(airline))
                if match and match.group(1).startswith("5"):
                    return "T1"
                return "T2"
            
            df2["Terminal"] = df2["Airline"].apply(get_terminal)
            
            # Step 3: Create "Profile" column (blank)
            df2["Profile"] = ""
            
            # Step 4: Create "Orig/Dest" column
            valid_cities = {"Karipur", "Singapore", "Jodhpur", "Cochin", "Phuket", "Goa", "Bangalore", "Mumbai", "Jabalpur", "Indore",
                            "Delhi", "Jaipur", "Allahabad", "Dubai", "Lucknow", "Hyderabad", "Ahmedabad", "Dhaka", "Vadodara", "Colombo",
                            "Patna", "Mangalore", "Madurai", "Udaipur", "Chennai", "Dehradun", "Abu Dhabi", "Bangkok", "Chandigarh", "Guwahati",
                            "Bhopal", "Ranchi", "Ayodhya", "Gwalior", "Srinagar", "Bahrain", "Rajkot", "Raipur", "Varanasi", "Calicut", "Kochi",
                            "New Delhi", "Kolkata", "Hyderabad", "Amritsar", "Thiruvananthapuram", "Agra", "Aurangabad", "Leh", "Kanpur", "Kandla",
                            "Jammu", "Imphal", "Belgaum", "Bhavnagar", "Bhuj", "Bhubaneswar", "Nagpur", "Gorakhpur", "Hubli", "Doha", "London",
                            "Zurich", "Hong Kong", "Kuwait", "New York", "Diu", "Male", "Dammam", "Mauritius", "Kuala Lumpur", "Prayagraj", "Toronto"}
            
            def clean_location(value):
                if pd.isna(value):
                    return ""
                for city in valid_cities:
                    if city in value:
                        return city
                return re.split(r"[,/]", value)[0].strip()
            
            df2["Origin"] = df2["Origin"].apply(clean_location)
            df2["Destination"] = df2["Destination"].apply(clean_location)
            
            def get_orig_dest(row):
                if row["ServiceType"] == "Arrival":
                    return row["Origin"]
                elif row["ServiceType"] == "Departure":
                    return row["Destination"]
                elif row["ServiceType"] == "Transit":
                    return f"{row['Origin']} / {row['Destination']}"
                return ""
            
            df2["Orig/Dest"] = df2.apply(get_orig_dest, axis=1)
            
            # Step 5: Create "ETA/ETD" column
            def calculate_eta_etd(row):
                if row["ServiceType"] == "Arrival":
                    return row["ServiceTime"]
                elif row["ServiceType"] == "Departure":
                    service_detail = str(row["ServiceDetail"]).lower()
                    if "domestic" in service_detail:
                        return (pd.to_datetime(row["ServiceTime"]) + pd.Timedelta(minutes=90)).time()
                    elif "international" in service_detail:
                        return (pd.to_datetime(row["ServiceTime"]) + pd.Timedelta(minutes=180)).time()
                return ""
            
            df2["ETA/ETD"] = df2.apply(calculate_eta_etd, axis=1)
            
            # Step 6: Create "Class of Travel" column (set to "Economy")
            df2["Class of Travel"] = "Economy"
            
            # Step 7: Create "Country Code" column (blank)
            df2["Country Code"] = ""
            
            # Step 8: Create "Nationality" column (blank)
            df2["Nationality"] = ""
            
            # Step 9: Create "Age" column (set to 30)
            df2["Age"] = 30
            
            # Step 10: Remove unwanted columns
            columns_to_remove = ["BookinDate", "FYMob", "DateOfService", "MOS", "ServiceLocation", "BookingStatus",
                                 "ArrivalFlightNumber", "DepartureFlightNumber", "FlightNumber", "Origin", "Destination",
                                 "PlacardGuestName", "PlacardCountryCode", "PlacardContactNo", "SupplierName", "SupplierGstNumber",
                                 "BookingGstCompanyName", "GstEmail", "PaymentMethod", "PaymentGateway", "PaymentMode",
                                 "CouponCode", "BaseFare", "Wsp", "DiscountAmount", "RevenueAmount", "TotalAmountPaid",
                                 "CreditCardTransactionNo", "UserRating", "Feedback", "InvoiceNumber", "QRScanDate",
                                 "QRScanTime", "SourceApp", "GstNumberCustomer", "GstCompanyName", "BillingEmail", "BillingAddress",
                                 "BillingCity", "BillingState", "BillingCountry", "BillingPincode", "SAPDocumentNumber",
                                 "SpecialInstruction", "SAPError"]
            
            df2.drop(columns=columns_to_remove, errors='ignore', inplace=True)
            
            # Step 11: Duplicate columns before renaming
            df2["GUEST NAME"] = df2["BookingPersonName"]
            df2["Placard Guest Name"] = df2["BookingPersonName"]
            df2["Placard Guest Contact No."] = df2["BookingPersonContactNo"]
            df2["Booker Contact No."] = df2["BookingPersonContactNo"]
            
            # Step 12: Rename columns
            rename_columns = {
                "ServiceTime": "Service Time", "ItenaryNumber": "Itinerary No.", "ServiceType": "Service",
                "ServiceDetail": "Package", "TotalPorters": "Total Guest",
                "Airline": "Flight No.", "BookingPersonEmail": "Email Id", "Remarks": "Remarks", "GSO": "GSO"
            }
            
            df2.rename(columns=rename_columns, inplace=True)
            
            # Step 13: Reorder columns
            column_order = ["Sr No", "Service Time", "Itinerary No.", "Service", "Package", "Terminal", "Profile", "GUEST NAME",
                            "Total Guest", "Flight No.", "Orig/Dest", "ETA/ETD", "Class of Travel", "Country Code",
                            "Placard Guest Contact No.", "Placard Guest Name", "Nationality", "Age", "Booker Contact No.", "Email Id",
                            "Remarks", "GSO"]
            
            df2 = df2.reindex(columns=column_order, fill_value="")
            
            # Step 15: Remove rows where "Remarks" is "Cancelled"
            df2 = df2[df2["Remarks"] != "Cancelled"]
            
            # Save the cleaned data to a new Excel file
            #df2.to_excel(output_file_path, index=False)
            

            
            def to_excel(df2):
                output = BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df2.to_excel(writer, index=False, sheet_name='Sheet1')
                processed_data = output.getvalue()
                return processed_data
            
            # Convert to Excel
            excel_data = to_excel(df2)
            
            
            #print("Data processing completed successfully! Output saved to:", output_file_path)
            #st.write("✅ Data processing completed successfully!")
            #st.write("### Processed Data:")
            #st.write(df2)
        

            # Step 17: Download Button
            #st.download_button(
                #label="Download Standalone File",
                #data=excel_data,
                #file_name="modified_data2.xlsx",
                #mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            #)
            
            
            
            
    if check_df is not None and check_df2 is not None:
        # Function to extract bold formatting from a column
        def extract_bold_styles(sheet, col_letter):
            bold_styles = {}
            for row in range(2, sheet.max_row + 1):  # Assuming first row is header
                cell = sheet[f"{col_letter}{row}"]
                bold_styles[row] = cell.font.bold if cell.font else False
            return bold_styles
        
        # Function to apply bold formatting
        def apply_bold_styles(sheet, col_letter, bold_styles):
            for row, is_bold in bold_styles.items():
                if is_bold:
                    sheet[f"{col_letter}{row}"].font = Font(bold=True)
        
        # Function to format headers
        def format_headers(sheet):
            for col in sheet.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
        
        # Function to process the two booking DataFrames
        def process_booking_dataframes(package_df, package_wb, standalone_df):
            # Load the original sheet to extract bold formatting
            package_ws = package_wb.active
            bold_styles = extract_bold_styles(package_ws, "E")  # Extracting bold from column E
        
            # Append standalone data to package data
            combined_df = pd.concat([package_df, standalone_df], ignore_index=True)
        
            # Sort by 'Service Time' if available
            if "Service Time" in combined_df.columns:
                combined_df = combined_df.sort_values(by=["Service Time"], ascending=True)
        
            # Reset 'Sr No.' to be continuous
            combined_df["Sr No."] = range(1, len(combined_df) + 1)
            st.write("### Final Processed DataFrame:")
            st.dataframe(combined_df)
        
            # Save to a BytesIO object using openpyxl to preserve formatting
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                combined_df.to_excel(writer, index=False, sheet_name="Sheet1")
        
            output.seek(0)  # Move cursor to the start
            output_wb = load_workbook(output)
            output_ws = output_wb.active
        
            # Apply the extracted bold styles back
            apply_bold_styles(output_ws, "E", bold_styles)
            format_headers(output_ws)  # Apply formatting to headers
        
            final_output = BytesIO()
            output_wb.save(final_output)
            final_output.seek(0)  # Move cursor to the start
        
            return final_output
        
        # Streamlit UI
        #st.title("Booking Data Processor")
        
        if excel_df and excel_data:
            # Read package DataFrame and preserve formatting
            excel_df.seek(0)  # Move cursor to the start
            package_wb = load_workbook(excel_df)
            package_df = pd.read_excel(excel_df, engine="openpyxl")
        
            # Read standalone DataFrame
            standalone_df = pd.read_excel(BytesIO(excel_data), engine="openpyxl")
        
            # Process the data while preserving formatting
            output = process_booking_dataframes(package_df, package_wb, standalone_df)
            #st.write(output)
        
            # Download Button
            st.download_button(
                label="Download Processed File",
                data=output,
                file_name="processed_booking.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


    
    

